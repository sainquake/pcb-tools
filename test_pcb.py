# test_capitalize.py
import os
import pytest
import json

def isclose(a, b, rel_tol=1e-03, abs_tol=0.0):
    return abs(a-b) <= max(rel_tol * max(abs(a), abs(b)), abs_tol)

def getProjectName(pa = './'):
    _PROJECT = None
    for path in os.listdir(pa):
        #print(path)
        if '.prjpcb' in path.lower():
            if '.PrjPcbStructure'.lower() not in path.lower():
                _PROJECT = path
                break
    return _PROJECT

def find_root_path_of_project():
    p = getProjectName()
    if p:
        return './'
    p = getProjectName('../../')
    if p:
        return '../../'
    p = getProjectName('../')
    if p:
        return '../'

def get_project_name():
    p = getProjectName()
    if p:
        return p.split('.')[0]
    p = getProjectName('../../')
    if p:
        return p.split('.')[0]
    p = getProjectName('../')
    if p:
        return p.split('.')[0]  

def test_root_path_found():
    """
    Root path of the project not found
    """
    p = find_root_path_of_project()
    assert p, test_root_path_found.__doc__


def check_files_exist():
    paths = []

    shouldBe = []
    shouldBe.append( {'root':'.','files':['.PrjPcb','.SchDoc','.BomDoc','.PCBDwf','LICENSE','.OutJob','.PcbDoc','.md','.gitignore'],'ok':False} )
    shouldBe.append( {'root':'doc','files':['.step','.pdf','view-bottom.png', 'view-top.png', 'view.png'],'ok':False} )
    shouldBe.append( {'root':'Outputs','files':['.Cam'],'ok':False} )
    shouldBe.append( {'root':'BOM','files':['BOM.xlsx'],'ok':False} )
    shouldBe.append( {'root':'Gerber','files':['.GBL','.GBO','.GBP','.GBS','.GTL','.GTO','.GTP','.GTS','.GM2'],'ok':False} )
    shouldBe.append( {'root':'NC Drill','files':['.txt'],'ok':False} )
    shouldBe.append( {'root':'Pick Place','files':['.csv','.txt'],'ok':False} )
    shouldBe.append( {'root':'Report Board Stack','files':['.xls'],'ok':False} )
    shouldBe.append( {'root':'WireListNetlist','files':['.net'],'ok':False} )
    
    for item in shouldBe:
        found_folder = False
        for root, subdirs, files in os.walk(find_root_path_of_project()):
    #print(root, subdirs, files)
            if ('.git' not in root) and ('History' not in root) and ('PCBWay-output' not in root):
            #print(root, subdirs, files)
            
                if (not item['ok']) and (item['root'] in root):
                    #print(files,item['files'])
                    for j in item['files']:
                        found = False
                        for k in files:
                            if j.lower() in k.lower():
                                found = True
                                #print(root,k,j)
                                paths.append( {'root':item['root'],'key':j,'path':root,'name':k} )
                                #print(paths[-1])
                                break 
                        if not found:
                            paths.append( {'root':item['root'],'key':j,'path':root,'name':False} )
                            print('----> ',print(paths[-1]))
                            #print(f'FILE {j} in {root} folder NOT FOUND')
                            #assert found, f'FILE {j} in {root} folder NOT FOUND, {test_files_exists.__doc__}'
                            #raise Exception(f'FILE {j} in {root} folder NOT FOUND')
                    item['ok'] = True
                    found_folder = True
                    break
        #print('--------------------',found_folder,item['root'])
        if found_folder:
            paths.append( {'root':item['root'],'key':'folder','path':root,'name':found_folder} )
            #print(paths[-1])
        else:
            paths.append( {'root':item['root'],'key':'folder','path':False,'name':found_folder} )
            print('---------->',paths[-1])
        
    return paths

# def find_extra_files():
#     paths = check_files_exist()
#     for item in paths:
#         print(item['path'],item['name'])

def test_files_exists():
    """
    This test checks existance of needed files.
    """
    paths = check_files_exist()
    for item in paths:
        assert item['name'], f'FOLDER { item["root"] } { item["key"] }, {test_files_exists.__doc__}'

def get_parameters_in_projpcb():
    paths = check_files_exist()
    prj_path = None
    for item in paths:
        if item['key'].lower() == '.PrjPcb'.lower():
            prj_path = item['path']+item['name']

    prj = open(prj_path,mode="r",encoding="utf-8")

    parameters = {}

    while prj:
        line = prj.readline()
        if '[Parameter' in line:
            n = prj.readline().split('=')[1].strip()
            v = prj.readline().split('=')[1].strip()
            parameters[n] = v
            print(n,v)
        if line == '':
            break
    prj.close()
    return parameters

def test_version_exists():
    """
    This test extract parameters from .prjpcb and check that Version parameter exists
    """
    parameters = get_parameters_in_projpcb()

    print(parameters)
    assert parameters['Version'], test_version_exists.__doc__


def generate_BOM():
    #EXTRACT Rv1 and Rv2 from BOM
    paths = check_files_exist()
    BOMPath = ''
    BOM = ''
    for item in paths:
        if 'BOM' in item['key']:
            print(item['name'])
            BOMPath = item['path'] +'/'
            BOM = item['path'] +'/'+item['name']
            break

    print (BOM)
    #if not os.path.isfile(BOM):
    #    print('BOM NOT FOUND')
        #raise Exception(f'BOM NOT FOUND')
    assert os.path.isfile(BOM), 'BOM NOT FOUND'

    BOMList = []
    try:
        from openpyxl import Workbook
        from openpyxl import load_workbook

        wb_obj = Workbook()
        wb_obj.template = True
        wb_obj = load_workbook(BOM) 
        wb_obj.template = True
        #
        sheet = wb_obj.active

        ns = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            #print(i,row)
            if i==0:
                ns = row
            else:
                li = {}
                for key,value in enumerate(row):
                    #print(key,value)
                    li[ns[key]] = value
                #print(li)
                BOMList.append(li)
    except:
        print('TRY BOM FROM TXT')

        assert os.path.isfile(BOMPath+'BOMtxt-BOM.txt'), 'BOM in TXT format NOT FOUND'
        # if not os.path.isfile(BOMPath+'BOMtxt-BOM.txt'):
        #     print('BOM in TXT format NOT FOUND')
        #     raise Exception(f'BOM in TXT format NOT FOUND')

        bom = open(BOMPath+'BOMtxt-BOM.txt',mode="r",encoding="ISO-8859-1")
        i=0
        ns = []
        while bom:
            line = bom.readline()
            if i==0:
                ns = line.split('\t')
            else:
                if len(line)>1:
                    row = line.split('\t')
                    li = {}
                    print(i,len(line),line.split('\t')[0])
                    for key,value in enumerate(row):
                        li[ns[key]] = value.replace('"','')
                    BOMList.append(li)
            i+=1
            if len(line)<1:
                break
        bom.close()

    #print(BOMList)
    return BOMList
    #writeData('BOMList.json',BOMList)


def test_bom_has_needed_keys():
    needed_keys = ['Designator','Quantity','MF','MP','Description','Value','Package','Type','Instructions']

    bom = generate_BOM()

    keys = (bom[0].keys())
    for item in needed_keys:
        print(item)
        assert (item in keys)

def extract_version_from_bom():
    bom = generate_BOM()
    Rv1_str = '0'
    Rv2_str = '0'

    for item in bom:
        #print(item['Designator'],item['Value'])
        if 'Rv1'.lower() in item['Designator'].lower():
            Rv1_str = item['Value']
        if 'Rv2'.lower() in item['Designator'].lower():
            Rv2_str = item['Value']

    Rv2 = float(Rv2_str.replace('k','').replace('M','')) * (1000 if 'k' in Rv2_str else 1) * (1000000 if 'M' in Rv2_str else 1)
    Rv1 = float(Rv1_str.replace('k','').replace('M','')) * (1000 if 'k' in Rv1_str else 1) * (1000000 if 'M' in Rv1_str else 1)

    #print('EXTRACT Rv1 and Rv2 from BOM ',Rv1,Rv2)
    return (Rv1,Rv2)

def extract_from_existed_boards_md():
    existed_boards_root = None
    for root, subdirs, files in os.walk(find_root_path_of_project()):
        if 'existed_boards.md' in files:
            #print(root, subdirs, files)
            existed_boards_root = root
            break
    
    assert (existed_boards_root), 'existed_boards.md not found'

    new_file=open(existed_boards_root+'/'+'existed_boards.md', mode="r", encoding="utf-8")
    Lines = new_file.readlines()  
    new_file.close()
    count = 0
    out = []
    for line in Lines:
        count += 1
        if count>2:
            s = line.strip().split('|')
            name = s[2].strip()
            r1 = float(s[3].strip() if len(s[3].strip())>1 else 0)
            r2 = float(s[4].strip() if len(s[4].strip())>1 else 0)
            #version in existed_boards
            v = str(s[5].strip()) 
            v_h = str(v.split('.')[0] if len(v.split('.'))>0 else 0)
            v_m = str(v.split('.')[1] if len(v.split('.'))>1 else 0)
            v_l = str(v.split('.')[2] if len(v.split('.'))>2 else 0)
            v = v_h+'.'+v_m+'.'+v_l
            out.append( {'name':name,'Rv1':r1,'Rv2':r2,'v':v,'v_h':v_h,'v_m':v_m,'v_l':v_l} )
            #print(name,r1,r2,v)
    return out

def test_version_in_existed_boards_md():
    parameters = get_parameters_in_projpcb()
    print(parameters['Version'])


    Rv1,Rv2 = extract_version_from_bom()

    print(f'BOM: EXTRACT Rv1 = {Rv1} and Rv2 = {Rv2}')

    eb = extract_from_existed_boards_md()

    pn = get_project_name()

    match = False
    name_match = False
    for item in eb:
        if (pn.lower() in item['name'].lower()):
                name_match = True
                if isclose(item['Rv1'] , Rv1) and isclose(item['Rv2'] , Rv2) and (item['v'] == parameters['Version']):
                    print('existed_boards.md:',item['name'],item['Rv1'],item['Rv2'],item['v'] ,'match!')
                    match = True
                else:
                    print('existed_boards.md:',item['name'],item['Rv1'],item['Rv2'],item['v'])
    
    if not match:
        print('No mached line in existed_boards.md: version is not found')
    if not name_match:
        print('No mached line in existed_boards.md: name is not found')

    assert name_match, 'No mached line in existed_boards.md: name is not found'
    assert match, 'No mached line in existed_boards.md: version is not found'

def test_wrong_MP():
    bom = generate_BOM()
    for item in bom:
        print(item['Designator'],item['Value'])
        
        if item['MP']!=None:
            assert not('502494-0670' in item['MP']), 'wrong MP, should be 502585-0670'
            assert not(('TJA1042' in item['MP']) and ('TK/3' not in item['MP'])), f'wrong MP, should be TJA1042TK/3'

            #if '502494-0670' in item['MP']:
                #raise Exception(f'Molex connector is wrong, 502494, should be 502585-0670')


            #if  ('TJA1042' in item['MP']) and ('TK/3' not in item['MP']):  
                #raise Exception(f'TJA1042 is wrong MP, should be TJA1042TK/3')


def generate_PNP():
    paths = check_files_exist()
    PNPPath = ''
    PNP = ''
    for item in paths:
        if type(item['name'])==str and 'Pick Place' in item['name'] and '.txt' in item['name']:
            print(item['name'])
            BOMPath = item['path'] +'/'
            PNP = item['path'] +'/'+item['name']
            break

    print (PNP)
    assert os.path.isfile(PNP), 'PNP NOT FOUND'

    pnp_data = []

    with open(PNP,mode="r",encoding="ISO-8859-1") as f:
        lines = f.readlines()

    start=False
    comment_index=0
    for item in lines:
        if start:
            pnp_data.append(item[0:comment_index].strip().replace('"',''))
        if 'Designator' in item:
            comment_index = item.index('Comment')
            start = True

    return pnp_data
    
def test_BOM_designators_in_PNP():
    bom = generate_BOM()
    pnp = generate_PNP()

    for item in bom:
        print(item['Designator'])
        des = item['Designator'].split(',')
        for it in des:
            assert it.strip() in pnp , f" {it} not in PNP"


if __name__ == "__main__":
    print(1)
    #(check_files_exist())
    #print(paths)

    #find_extra_files()

    #PROJECT = getProjectName()
    #test_version_exists()
    #print(PROJECT)
    #test_bom_has_needed_keys()
    #print(find_root_path_of_project())

    # bom = generate_BOM()
    # for item in bom:
    #     print(item['Designator'])
    #test_version_in_existed_boards_md()



