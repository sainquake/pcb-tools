

def alphaAndCut(input_path,output_path):
    from rembg import remove
    import numpy as np
    import cv2
    src = cv2.imread(input_path)
    tgt = remove(src)
    rows = np.any(tgt, axis=1)
    cols = np.any(tgt, axis=0)
    rmin, rmax = np.where(rows)[0][[0, -1]]
    cmin, cmax = np.where(cols)[0][[0, -1]]
    cv2.imwrite(output_path, tgt[rmin:rmax, cmin:cmax])

def getBOM():
    import pandas as pd
    with open('./Project Outputs/BOM/BOMtxt-BOM.txt') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    for i in range(lines.__len__()):
        lines[i] = lines[i].split('\t')
    for i in range(len(lines)):
        for j in range(len(lines[i])):
            lines[i][j] = lines[i][j].replace('"', '')
    data = pd.DataFrame(lines[1:], columns=lines[0])
    data = data.drop(0)

    return data

def generatBOMToPCBWay(save_to = './Project Outputs/BOM/output.xlsx'):
    import openpyxl
    import pandas as pd
    data = getBOM()
    columns = ['*Designator', '*Qty', 'Manufacturer', '*Mfg Part #', 'Value / Description', '*Package/Footprint',
            'Mounting Type', 'Your Instructions / Notes', 'Assembly', '*Unit Price(XX sets)', '*Total', '*Delivery Time',
            '*Actual Purchase Mfg Part #', '*PCBWay Note', 'Customer Reply', 'PCBWay Update']
    dataxls = pd.DataFrame(columns=columns)
    dataxls['*Designator'] = data['Designator']
    dataxls['*Qty'] = data['Quantity']
    dataxls['Manufacturer'] = data['MF']
    dataxls['*Mfg Part #'] = data['MP']
    dataxls['Value / Description'] = data['Value'].astype('str') + '; ' + data['Description'].astype('str')
    dataxls['*Package/Footprint'] = data['Package']
    dataxls['Mounting Type'] = data['Type']
    dataxls['Your Instructions / Notes'] = data['Instructions'] + '; ' + data['HelpURL']
    dataxls.to_excel(save_to)

    book = openpyxl.load_workbook(save_to)
    sheet = book['Sheet1']
    for col in sheet.columns: # автоматическая ширина ячеек
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length * 0.8
        sheet.column_dimensions[column].width = adjusted_width if adjusted_width<50 else 50
    yellow = openpyxl.styles.PatternFill('solid', fgColor="FFFF00")
    grey = openpyxl.styles.PatternFill('solid', fgColor="C0C0C0")
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = grey
    sheet['B1'].fill = yellow
    sheet['C1'].fill = yellow
    sheet['E1'].fill = yellow
    sheet['G1'].fill = yellow
    sheet['K1'].fill = yellow
    sheet['L1'].fill = yellow
    sheet['M1'].fill = yellow
    sheet['N1'].fill = yellow
    sheet['O1'].fill = yellow
    sheet['P1'].fill = yellow
    sheet['Q1'].fill = yellow
    book.save(save_to)
    return data



def parseNetlist():
    import os
    #if not os.path.isfile(path)
    net = os.listdir('./Project Outputs/WireListNetlist/')[0]

    with open(f'./Project Outputs/WireListNetlist/{net}') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    lines = lines[lines.index('<<< Wire List >>>'):]
    while '' in lines:
        lines.remove('')

    keys = ['net','designator','pinNum','pinName','component']
    data = lines[2:]
    for i in range(data.__len__()):
        data[i] = data[i].split()
    result = []
    net = ''
    for i in data:
        if i[0][0] == '[':
            net = i[1]
        else:
            result.append({
                'net':net,
                'designator': i[0],
                'pinNum': i[1],
                'pinName': i[2],
                'component': i[-1],
                })
    return result


def getBBoxFromSTEP():
    # pip install steputils
    import os
    from steputils import p21
    step_file = ''

    for path in os.listdir('./doc'):
        if '.step' in path.lower():
            step_file = path
            break
    
    file = p21.readfile(f'doc/{step_file}')

    points = [
        file.data[0].instances[x].entity.params[1] for x in [                      # 4 get point by id
            entry.entity.params[1] for entry in                                    # 3 get vertex point id
            sum([list(section.instances.values()) for section in file.data], [])   # 1 gather all sections
            if hasattr(entry, 'entity') and entry.entity.name == 'VERTEX_POINT'    # 2 filter all vertices
        ]
    ]

    min_x = min(points, key=lambda x: x[0])[0]
    max_x = max(points, key=lambda x: x[0])[0]

    min_y = min(points, key=lambda x: x[1])[1]
    max_y = max(points, key=lambda x: x[1])[1]

    min_z = min(points, key=lambda x: x[2])[2]
    max_z = max(points, key=lambda x: x[2])[2]

    bbox = (
        (min_x, min_y, min_z),
        (max_x, max_y, max_z)
    )

    dim = (
        max_x - min_x,
        max_y - min_y,
        max_z - min_z
    )

    #print(bbox)
    #print(dim)

    x = round(dim[0]*10)/10.0
    y = round(dim[1]*10)/10.0
    z = round(dim[2]*10)/10.0

    print(x,y,z)
    return [x,y,z]

def getBBoxFromGerber(file_name='PCB.GM2'):
    from gerber import load_layer
    # Open the gerber files
    gm2 = load_layer(f'./Project Outputs/Gerber/{file_name}')

    #print(gm2.__dict__)
    #print(gm2.bounds)
    print(-gm2.bounds[0][0]+gm2.bounds[0][1])
    print(-gm2.bounds[1][0]+gm2.bounds[1][1])
    gerber_x = round((-gm2.bounds[0][0]+gm2.bounds[0][1])*100)/100
    gerber_y = round((-gm2.bounds[1][0]+gm2.bounds[1][1])*100)/100
    #gm2.close()
    return [gerber_x,gerber_y]


def layerStackParce():
    import pandas as pd
    data = pd.read_excel('./Project Outputs/Report Board Stack/PCB.xls')
    # суммарная толщина всех слоёв
    maxvalue = float(list(data[-1:].stack())[0].split()[-1].replace('mm','').replace(',','.')) # Последняя строка из таблицы стакается в одну ячейку, превращается в список, оставшаяся строка делится по пробелам, удаляется mm, запятая меняется на точку, всё это превращается в число float
    data = data[4:-3]
    layers = data.loc[data.iloc[:,-3] == 'Copper'] #Ищу по столбцу где есть медь - это проводящие слои
    listlayers = list(layers.iloc[:,-4]) # Список слоёв
    hight = list(layers.iloc[:,-2]) # Список толщин
    temp = []
    for i in hight:
        temp.append(float(i.replace('mm','').replace(',','.')))
    hight = temp.copy()
    standrdrow = [0.2, 0.3, 0.4, 0.6, 0.8, 1.0, 1.2, 1.6, 2.0, 2.4, 2.6, 2.8, 3.0, 3.2]
    standrdvalue = [abs(i-maxvalue) for i in standrdrow]
    standrdvalue = standrdrow[standrdvalue.index(min(standrdvalue))]

    re = {}
    re['height'] = hight
    re['list'] = listlayers
    re['stndart height'] = standrdvalue
    return re

def minTrace(file_name = 'PCB.GTL'):
    from gerber import load_layer
    with open(f'./Project Outputs/Gerber/{file_name}') as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip('*\n')
    listfromPCBG1 = []
    for i in lines:
        if i.find('%') == False: #TODO поиск строк по % во всём файле вообще корректен?
            i = i.strip('*%')
            if i.find('C') != -1 and i.find(',') != -1:
                listfromPCBG1.append(float(i[i.find(',')+1:]))

    return min(listfromPCBG1)

def minFromNCDrill(file_name = './Project Outputs/NC Drill/PCB.TXT'):
    with open(file_name) as f:
        lines = f.readlines()
    for i in range(lines.__len__()):
        lines[i] = lines[i].strip()
    pcbplate = lines[lines.index(';TYPE=PLATED')+1:lines.index('%')] #TODO в цикл его надо пихать шобы по всему файлу пройтись?
    tmp = []
    for i in pcbplate:
        if 'F00S00C' in i:
            tmp.append(float(i[i.index('C')+1:]))
    return min(tmp)

def minFromAllNCDrill():
    import os
    drill_files = []
    for path in os.listdir('./Project Outputs/NC Drill/'):
        if '.TXT'.lower() in path.lower():
            drill_files.append('./Project Outputs/NC Drill/'+path)
            #shutil.copyfile('./Project Outputs/NC Drill/'+path, './PCBWay-output/'+path)

    mindrill = 50
    for item in drill_files:
        mindrill = min( minFromNCDrill(item),mindrill)